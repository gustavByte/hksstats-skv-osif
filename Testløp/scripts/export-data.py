from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "data" / "Testlop_HKS.xlsx"
EXTRA_2026_WORKBOOK_PATH = ROOT / "data" / "2026_HKS_Testlop_Resultater_Pamelding_Hjelpere.xlsx"
OUTPUT_DIR = ROOT / "src" / "data"

DISTANCES = (600, 1200)
GENDERS = ("Kvinner", "Menn")
APPROVED_CHECK_RESULT_IDS = {
    "R0035",  # Sofie Kjos Bergum, 600 m, 2022
    "R0072",  # Elisabeth Oiulfstad, 1200 m, 2023
    "R0045",  # Vilde Antonsen, 600 m, 2023
    "R0350",  # Dina Anjara H. R., 600 m, 2025
    "R0107",  # Tea Saetereng Fyksen, 1200 m, 2023
    "R0074",  # Rune Solli, 1200 m, 2023
    "R0138",  # Julie Kvale Stostad, 1200 m, 2023
    "R0140",  # Hannah Engeskaug Nilsen, 1200 m, 2023
    "R0143",  # Erik Ronneberg, 1200 m, 2023
    "R0152",  # Johanne Laegran, 1200 m, 2023
    "R0134",  # Veslemoy Ronnevik, 600 m, 2023
    "R0230",  # Mari Brondbo Dahl, 1200 m, 2024
    "R0251",  # Tina Kleven, 600 m, 2024
    "R0269",  # Bastian Fjermeros, 1200 m, 2024
}
EXCLUDED_RESULT_IDS = {
    "R0204",  # Paameldt, lop ikke
    "R0205",  # Paameldt, lop ikke
    "R0206",  # Paameldt, lop ikke
    "R0207",  # Paameldt, lop ikke
}
NAME_ALIAS_OVERRIDES = {
    "Anne-Line Evenstad Dalen": "Anne-Line Evenstad Dahlen",
    "Martin Stay Egeberg": "Martin Stray Egeberg",
    "Tobias Bauman": "Tobias Baumann",
    "Frdrik Fyksen": "Fredrik Sætereng Fyksen",
    "Henning Mæhle": "Henning Mæle",
    "Hanne Tingstad": "Hanne Tingelstad",
    "Erlend G. Sørtveit": "Erlend Gjerdevik Sørtveit",
    "Erlend Sørtveit": "Erlend Gjerdevik Sørtveit",
    "Telma Dagestad": "Telma Eckhoff Dagestad",
    "Veslemøy": "Veslemøy Rønnevik",
    "Anna Leikanger Aasen": "Anna Liv Leikanger Aasen",
    "Henrik Hansen": "Henrik Victor Hansen",
    "Ida Nordengen": "Ida Marie Sirevaag Nordengen",
    "Maja Blegen": "Maja Maria Blegen",
    "Sofie Amundsgård": "Sofie Karlsen Amundsgård",
    "Sofie Kjos": "Sofie Kjos Bergum",
    "Anne Cecilie Grindheim": "Anne Cecilie Grindstad",
    "Fredrik Fyksen": "Fredrik Sætereng Fyksen",
    "Matilde Belsvik": "Matilde Reinholdt Belsvik",
    "Siri Staver": "Siri Vegsund Staver",
    "Elisavbeth Ø": "Elisabeth Øiulfstad",
}

MANUAL_NAME_REVIEW_GROUPS = [
    ["Anne-Line Evenstad Dahlen", "Anne-Line Evenstad Dalen"],
    ["Martin Stay Egeberg", "Martin Stray Egeberg"],
    ["Tobias Bauman", "Tobias Baumann"],
    ["Frdrik Fyksen", "Fredrik Fyksen"],
    ["Henning Mæhle", "Henning Mæle"],
    ["Hanne Tingelstad", "Hanne Tingstad"],
    ["Erlend G. Sørtveit", "Erlend Sørtveit", "Erlend Gjerdevik Sørtveit"],
    ["Telma Dagestad", "Telma Eckhoff Dagestad"],
    ["Veslemøy", "Veslemøy Rønnevik"],
    ["Anna Leikanger Aasen", "Anna Liv Leikanger Aasen"],
    ["Henrik Hansen", "Henrik Victor Hansen"],
    ["Ida Nordengen", "Ida Marie Sirevaag Nordengen"],
    ["Maja Blegen", "Maja Maria Blegen"],
    ["Sofie Amundsgård", "Sofie Karlsen Amundsgård"],
    ["Sofie Kjos", "Sofie Kjos Bergum"],
    ["Anne Cecilie Grindheim", "Anne Cecilie Grindstad"],
    ["Fredrik Fyksen", "Fredrik Sætereng Fyksen"],
    ["Matilde Belsvik", "Matilde Reinholdt Belsvik"],
    ["Siri Staver", "Siri Vegsund Staver"],
    ["Elisavbeth Ø", "Elisabeth Øiulfstad"],
]


def slugify(value: str) -> str:
    value = (
        value.replace("Æ", "Ae")
        .replace("æ", "ae")
        .replace("Ø", "O")
        .replace("ø", "o")
        .replace("Å", "A")
        .replace("å", "a")
    )
    stripped = "".join(
        ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch)
    )
    stripped = stripped.lower().strip()
    stripped = re.sub(r"[^a-z0-9]+", "_", stripped)
    return re.sub(r"_+", "_", stripped).strip("_")


def canonical_name(value: str) -> str:
    return NAME_ALIAS_OVERRIDES.get(value, value)


def canonical_person_id(value: str) -> str:
    return slugify(canonical_name(value))


def as_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def as_bool(value: Any) -> bool:
    return str(value).strip().lower() in {"ja", "true", "1", "yes"}


def as_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def as_iso_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except (TypeError, ValueError):
            return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def parse_time_seconds(value: Any) -> float | None:
    text = as_text(value)
    if not text:
        return None
    cleaned = text.replace(",", ".")
    if ":" in cleaned:
        minute_text, second_text = cleaned.split(":", 1)
        try:
            return round(int(minute_text) * 60 + float(second_text), 2)
        except ValueError:
            return None
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return None


def format_time_display(value: Any) -> str | None:
    text = as_text(value)
    if not text:
        return None
    return text.replace(".", ",")


def sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def result_from_row(row: dict[str, Any]) -> dict[str, Any]:
    raw_name = as_text(row.get("navn_normalisert")) or as_text(row.get("navn_original")) or "Ukjent"
    name = canonical_name(raw_name)
    person_id = (
        canonical_person_id(raw_name)
        if raw_name != name or name in NAME_ALIAS_OVERRIDES.values()
        else as_text(row.get("person_id")) or slugify(name)
    )
    result = {
        "id": as_text(row.get("resultat_id")),
        "year": as_int(row.get("år")),
        "testlopId": as_text(row.get("testløp_id")),
        "date": as_iso_date(row.get("dato")),
        "place": as_text(row.get("sted")),
        "distance": as_int(row.get("distanse_m")),
        "gender": as_text(row.get("kjønn")),
        "name": name,
        "originalName": as_text(row.get("navn_original")),
        "personId": person_id,
        "timeOriginal": as_text(row.get("tid_original")),
        "timeSeconds": as_number(row.get("tid_sekunder")),
        "timeDisplay": as_text(row.get("tid_visning")),
        "note": as_text(row.get("merknad")),
        "validToplist": as_bool(row.get("gyldig_toppliste")),
        "validRecord": as_bool(row.get("gyldig_rekord")),
        "sourceSheet": as_text(row.get("kildeark")),
        "sourceCell": as_text(row.get("kildecelle")),
        "sourceType": as_text(row.get("kildetype")),
        "checkStatus": as_text(row.get("sjekk_status")) or "OK",
    }
    return apply_approved_check(result)


def build_2026_results() -> list[dict[str, Any]]:
    if not EXTRA_2026_WORKBOOK_PATH.exists():
        return []

    workbook = load_workbook(EXTRA_2026_WORKBOOK_PATH, read_only=True, data_only=True)
    results: list[dict[str, Any]] = []
    result_index = 1

    for sheet in workbook.worksheets:
        if not sheet.title.startswith("HKS 2026"):
            continue
        date_match = re.search(r"(2026-\d{2}-\d{2})", sheet.title)
        lop_match = re.search(r"Løp #(\d+)", sheet.title)
        if not date_match or not lop_match:
            continue

        testlop_id = f"2026_HKS_{lop_match.group(1)}"
        sections: dict[int, dict[str, Any]] = {}
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = list(row[:4]) + [None] * 4
            left_name, left_time, right_name, right_time = values[:4]

            for column, value in ((0, left_name), (2, right_name)):
                if isinstance(value, str) and ("600" in value or "1200" in value):
                    sections[column] = {
                        "distance": 600 if "600" in value else 1200,
                        "gender": "Kvinner" if "Kvinner" in value else "Menn",
                    }

            for column, name, time_value in (
                (0, left_name, left_time),
                (2, right_name, right_time),
            ):
                if column not in sections:
                    continue
                original_name = as_text(name)
                time_seconds = parse_time_seconds(time_value)
                if not original_name or original_name.lower() == "navn" or time_seconds is None:
                    continue

                cell_range = f"{'A' if column == 0 else 'C'}{row_index}:{'B' if column == 0 else 'D'}{row_index}"
                row_data = {
                    "resultat_id": f"R2026_{result_index:03d}",
                    "år": 2026,
                    "testløp_id": testlop_id,
                    "dato": date_match.group(1),
                    "sted": None,
                    "distanse_m": sections[column]["distance"],
                    "kjønn": sections[column]["gender"],
                    "navn_original": original_name,
                    "navn_normalisert": original_name,
                    "tid_original": as_text(time_value),
                    "tid_sekunder": time_seconds,
                    "tid_visning": format_time_display(time_value),
                    "merknad": None,
                    "gyldig_toppliste": "ja",
                    "gyldig_rekord": "ja",
                    "kildeark": sheet.title,
                    "kildecelle": cell_range,
                    "kildetype": "2026-arbeidsbok",
                    "sjekk_status": "OK",
                }
                results.append(result_from_row(row_data))
                result_index += 1

    return results


def apply_approved_check(result: dict[str, Any]) -> dict[str, Any]:
    if result["id"] in APPROVED_CHECK_RESULT_IDS:
        result["validToplist"] = True
        result["validRecord"] = True
        result["checkStatus"] = "OK"
    return result


def public_result(result: dict[str, Any]) -> dict[str, Any]:
    return dict(result)


def sort_results(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda item: (
            item["year"] or 0,
            item["date"] or "",
            item["distance"] or 0,
            item["gender"] or "",
            item["timeSeconds"] if item["timeSeconds"] is not None else 99999,
            item["name"],
        ),
    )


def fastest_per_person(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    best: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row["timeSeconds"] is None:
            continue
        current = best.get(row["personId"])
        if current is None or row["timeSeconds"] < current["timeSeconds"]:
            best[row["personId"]] = row
    return sorted(
        best.values(),
        key=lambda item: (
            item["timeSeconds"] if item["timeSeconds"] is not None else 99999,
            item["date"] or "",
            item["name"],
        ),
    )


def rank_entries(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ranked = []
    for index, row in enumerate(rows, start=1):
        entry = public_result(row)
        entry["rank"] = index
        ranked.append(entry)
    return ranked


def build_people(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for result in results:
        grouped[result["personId"]].append(result)

    people: list[dict[str, Any]] = []
    for person_id, rows in grouped.items():
        name = Counter(row["name"] for row in rows).most_common(1)[0][0]
        years = sorted({row["year"] for row in rows if row["year"]})
        valid_rows = [
            row for row in rows if row["validToplist"] and row["timeSeconds"] is not None
        ]
        pbs = {}
        for distance in DISTANCES:
            candidates = [row for row in valid_rows if row["distance"] == distance]
            best = min(candidates, key=lambda row: row["timeSeconds"], default=None)
            pbs[str(distance)] = public_result(best) if best else None
        people.append(
            {
                "personId": person_id,
                "name": name,
                "resultCount": len(rows),
                "firstYear": min(years) if years else None,
                "lastYear": max(years) if years else None,
                "years": years,
                "genders": sorted({row["gender"] for row in rows if row["gender"]}),
                "distances": sorted({row["distance"] for row in rows if row["distance"]}),
                "pb": pbs,
            }
        )
    return sorted(people, key=lambda item: item["name"])


def build_leaderboards(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    leaderboards: list[dict[str, Any]] = []
    for distance in DISTANCES:
        for gender in GENDERS:
            official_candidates = [
                row
                for row in results
                if row["distance"] == distance
                and row["gender"] == gender
                and row["validToplist"]
                and row["timeSeconds"] is not None
            ]
            with_checks_candidates = [
                row
                for row in results
                if row["distance"] == distance
                and row["gender"] == gender
                and row["timeSeconds"] is not None
            ]
            leaderboards.append(
                {
                    "distance": distance,
                    "gender": gender,
                    "official": rank_entries(fastest_per_person(official_candidates)),
                    "withChecks": rank_entries(fastest_per_person(with_checks_candidates)),
                }
            )
    return leaderboards


def build_records(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for distance in DISTANCES:
        for gender in GENDERS:
            official = min(
                (
                    row
                    for row in results
                    if row["distance"] == distance
                    and row["gender"] == gender
                    and row["validRecord"]
                    and row["timeSeconds"] is not None
                ),
                key=lambda row: (row["timeSeconds"], row["date"] or "", row["name"]),
                default=None,
            )
            pending = min(
                (
                    row
                    for row in results
                    if row["distance"] == distance
                    and row["gender"] == gender
                    and row["timeSeconds"] is not None
                    and (
                        row["checkStatus"] != "OK"
                        or not row["validRecord"]
                        or not row["validToplist"]
                    )
                ),
                key=lambda row: (row["timeSeconds"], row["date"] or "", row["name"]),
                default=None,
            )
            records.append(
                {
                    "distance": distance,
                    "gender": gender,
                    "official": public_result(official) if official else None,
                    "pending": public_result(pending) if pending else None,
                }
            )
    return records


def build_matrices(results: list[dict[str, Any]], years: list[int]) -> list[dict[str, Any]]:
    matrices: list[dict[str, Any]] = []
    for distance in DISTANCES:
        for gender in GENDERS:
            rows = [
                row
                for row in results
                if row["distance"] == distance
                and row["gender"] == gender
                and row["timeSeconds"] is not None
            ]
            by_person_year: dict[tuple[str, int], dict[str, Any]] = {}
            for row in rows:
                key = (row["personId"], row["year"])
                current = by_person_year.get(key)
                if current is None or row["timeSeconds"] < current["timeSeconds"]:
                    by_person_year[key] = row
            person_ids = sorted({row["personId"] for row in rows})
            person_rows = []
            for person_id in person_ids:
                person_results = [row for row in rows if row["personId"] == person_id]
                name = Counter(row["name"] for row in person_results).most_common(1)[0][0]
                pb = min(
                    (row for row in person_results if row["validToplist"]),
                    key=lambda row: row["timeSeconds"],
                    default=None,
                )
                cells = {
                    str(year): public_result(by_person_year[(person_id, year)])
                    for year in years
                    if (person_id, year) in by_person_year
                }
                best_sort = pb["timeSeconds"] if pb else min(row["timeSeconds"] for row in person_results)
                person_rows.append(
                    {
                        "personId": person_id,
                        "name": name,
                        "cells": cells,
                        "pb": public_result(pb) if pb else None,
                        "sortSeconds": best_sort,
                    }
                )
            matrices.append(
                {
                    "distance": distance,
                    "gender": gender,
                    "years": years,
                    "rows": sorted(person_rows, key=lambda item: (item["sortSeconds"], item["name"])),
                }
            )
    return matrices


def build_stats(results: list[dict[str, Any]], years: list[int]) -> dict[str, Any]:
    year_stats = []
    for year in years:
        rows = [row for row in results if row["year"] == year]
        breakdown = {
            f"{distance}_{gender}": sum(
                1
                for row in rows
                if row["distance"] == distance and row["gender"] == gender
            )
            for distance in DISTANCES
            for gender in GENDERS
        }
        year_stats.append(
            {
                "year": year,
                "resultCount": len(rows),
                "personCount": len({row["personId"] for row in rows}),
                "testlopCount": len({row["testlopId"] for row in rows if row["testlopId"]}),
                "deviationCount": sum(1 for row in rows if row["checkStatus"] != "OK"),
                "breakdown": breakdown,
            }
        )

    testlop_stats = []
    for testlop_id in sorted({row["testlopId"] for row in results if row["testlopId"]}):
        rows = [row for row in results if row["testlopId"] == testlop_id]
        dates = sorted({row["date"] for row in rows if row["date"]})
        places = [row["place"] for row in rows if row["place"]]
        testlop_stats.append(
            {
                "testlopId": testlop_id,
                "year": rows[0]["year"],
                "date": dates[0] if dates else None,
                "place": Counter(places).most_common(1)[0][0] if places else None,
                "resultCount": len(rows),
                "personCount": len({row["personId"] for row in rows}),
                "distance600": sum(1 for row in rows if row["distance"] == 600),
                "distance1200": sum(1 for row in rows if row["distance"] == 1200),
                "women": sum(1 for row in rows if row["gender"] == "Kvinner"),
                "men": sum(1 for row in rows if row["gender"] == "Menn"),
                "deviationCount": sum(1 for row in rows if row["checkStatus"] != "OK"),
            }
        )

    result_years = sorted({row["year"] for row in results if row["year"]}, reverse=True)
    latest_year = max(
        (stat["year"] for stat in year_stats if stat["resultCount"] > 0),
        default=None,
    )
    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceWorkbook": str(WORKBOOK_PATH.relative_to(ROOT)).replace("\\", "/"),
        "years": years,
        "resultYears": result_years,
        "latestYear": latest_year,
        "totals": {
            "results": len(results),
            "people": len({row["personId"] for row in results}),
            "testlop": len({row["testlopId"] for row in results if row["testlopId"]}),
            "deviations": sum(1 for row in results if row["checkStatus"] != "OK"),
        },
        "yearStats": sorted(year_stats, key=lambda item: item["year"], reverse=True),
        "testlopStats": sorted(
            testlop_stats,
            key=lambda item: (item["year"] or 0, item["date"] or "", item["testlopId"]),
            reverse=True,
        ),
    }


def build_quality(
    workbook: Any,
    results: list[dict[str, Any]],
    aliases: list[dict[str, Any]],
) -> dict[str, Any]:
    deviations = []
    if "Avvik og sjekk" in workbook.sheetnames:
        for row in sheet_rows(workbook, "Avvik og sjekk"):
            item = result_from_row(row)
            if item["id"] in APPROVED_CHECK_RESULT_IDS or item["id"] in EXCLUDED_RESULT_IDS:
                continue
            item["checkNote"] = as_text(row.get("sjekk_merknad"))
            deviations.append(item)

    alias_by_name = {
        as_text(row.get("alias")): {
            "alias": as_text(row.get("alias")),
            "personId": canonical_person_id(as_text(row.get("alias")))
            if as_text(row.get("alias")) in NAME_ALIAS_OVERRIDES
            else as_text(row.get("person_id")),
            "status": "Avklart i nettside"
            if as_text(row.get("alias")) in NAME_ALIAS_OVERRIDES
            else as_text(row.get("status")),
            "comment": as_text(row.get("kommentar")),
        }
        for row in aliases
    }
    manual_name_review = []
    archived_name_review = []
    for group in MANUAL_NAME_REVIEW_GROUPS:
        entries = [
            alias_by_name.get(
                name,
                {
                    "alias": name,
                    "personId": canonical_person_id(name),
                    "status": "Avklart i nettside",
                    "comment": None,
                },
            )
            for name in group
        ]
        person_ids = sorted({canonical_person_id(name) for name in group})
        review = {
            "names": group,
            "entries": entries,
            "personIds": person_ids,
            "needsReview": len(person_ids) != 1,
        }
        if review["needsReview"]:
            manual_name_review.append(review)
        else:
            archived_name_review.append(review)

    return {
        "deviations": deviations,
        "manualNameReview": manual_name_review,
        "archivedNameReview": archived_name_review,
        "summary": {
            "deviationCount": len(deviations),
            "manualNameReviewCount": len(manual_name_review),
            "archivedNameReviewCount": len(archived_name_review),
        },
    }


def infer_years(workbook: Any, results: list[dict[str, Any]]) -> list[int]:
    years = {row["year"] for row in results if row["year"]}
    for sheet_name in workbook.sheetnames:
        for match in re.findall(r"20\d{2}", sheet_name):
            years.add(int(match))
    return sorted(years, reverse=True)


def write_json(filename: str, payload: Any) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / filename
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Fant ikke {WORKBOOK_PATH}")

    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    raw_results = sheet_rows(workbook, "Resultater")
    raw_aliases = sheet_rows(workbook, "Navnealias") if "Navnealias" in workbook.sheetnames else []

    results = []
    for row in raw_results:
        result = result_from_row(row)
        if result["id"] not in EXCLUDED_RESULT_IDS:
            results.append(result)
    results.extend(build_2026_results())
    results = sort_results(results)
    people = build_people(results)
    years = infer_years(workbook, results)
    stats = build_stats(results, years)
    leaderboards = build_leaderboards(results)
    records = build_records(results)
    matrices = build_matrices(results, years)
    quality = build_quality(workbook, results, raw_aliases)

    write_json("results.json", results)
    write_json("people.json", people)
    write_json("stats.json", stats)
    write_json("leaderboards.json", leaderboards)
    write_json("records.json", records)
    write_json("matrices.json", matrices)
    write_json("quality.json", quality)

    print(
        f"Exported {len(results)} results, {len(people)} people, "
        f"{len(leaderboards)} leaderboards and {len(quality['deviations'])} deviations."
    )


if __name__ == "__main__":
    main()
