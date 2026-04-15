from __future__ import annotations

import csv
import json
import re
import shutil
import sqlite3
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
DOWNLOADS_DIR = ROOT / "public" / "downloads"
PUBLIC_DATA_DIR = ROOT / "public" / "data"
DB_PATH = DATA_DIR / "hksstats.sqlite"
REVIEW_PATH = DATA_DIR / "name_match_review.csv"


CLASS_META = {
    "EliteSKV": {
        "label": "SK Vidar elite",
        "organization": "SKV",
        "official_label": "Elite",
        "sort_order": 1,
    },
    "SeniorSKV": {
        "label": "SK Vidar senior",
        "organization": "SKV",
        "official_label": "Senior",
        "sort_order": 2,
    },
    "Veteran": {
        "label": "SK Vidar veteran",
        "organization": "SKV",
        "official_label": "Veteran",
        "sort_order": 3,
    },
    "StudOSI": {
        "label": "OSI student",
        "organization": "OSIF",
        "official_label": "Student",
        "sort_order": 4,
    },
    "MiksSKV": {
        "label": "SK Vidar miks",
        "organization": "SKV",
        "official_label": "A3 Lag tilsluttet andre særforbund i NIF",
        "sort_order": 5,
    },
    "MiksOSI": {
        "label": "OSI miks",
        "organization": "OSIF",
        "official_label": "A3 Lag tilsluttet andre særforbund i NIF",
        "sort_order": 6,
    },
}

ORGANIZATIONS = {
    "SKV": {"name": "SK Vidar", "short_name": "SKV"},
    "OSIF": {"name": "OSI Friidrett", "short_name": "OSIF"},
}

STANDARD_STAGE_DISTANCES_M = {
    1: 1170,
    2: 1130,
    3: 595,
    4: 1920,
    5: 1210,
    6: 1250,
    7: 1770,
    8: 1780,
    9: 625,
    10: 2860,
    11: 1520,
    12: 350,
    13: 1080,
    14: 710,
    15: 535,
}

STAGE_DISTANCE_SOURCE = (
    "https://holmenkollstafetten.no/nyheter/om-m%C3%A5ling-av-etappenes-distanser"
)


@dataclass
class TeamRecord:
    year: int
    organization_code: str
    class_code: str
    team_name: str
    source_sheet: str
    header_row: int
    group_index: int
    total_time_text: str | None
    total_seconds: int | None
    team_rank: int | None


@dataclass
class ResultRecord:
    year: int
    organization_code: str
    class_code: str
    team_name: str
    source_sheet: str
    source_row: int
    source_col: int
    header_row: int
    group_index: int
    division: str
    stage_number: int
    stage_label: str
    stage_label_source: str
    raw_name: str
    split_text: str | None
    split_seconds: int | None
    oa_rank: int | None
    category_rank: int | None


def ensure_directories() -> None:
    for directory in (DATA_DIR, DOWNLOADS_DIR, PUBLIC_DATA_DIR):
        directory.mkdir(parents=True, exist_ok=True)


def find_workbook() -> Path:
    workbooks = sorted(ROOT.glob("*.xlsx"))
    if not workbooks:
        raise FileNotFoundError("Fant ingen .xlsx-fil i prosjektroten.")
    return workbooks[0]


def normalize_name(value: str) -> str:
    stripped = "".join(
        ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch)
    )
    stripped = stripped.lower().strip()
    stripped = re.sub(r"[^a-z0-9 ]+", " ", stripped)
    return re.sub(r"\s+", " ", stripped).strip()


def tokenize_name(value: str) -> list[str]:
    return normalize_name(value).split()


def edit_distance(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    previous = list(range(len(b) + 1))
    for i, ca in enumerate(a, start=1):
        current = [i]
        for j, cb in enumerate(b, start=1):
            insert_cost = current[j - 1] + 1
            delete_cost = previous[j] + 1
            replace_cost = previous[j - 1] + (ca != cb)
            current.append(min(insert_cost, delete_cost, replace_cost))
        previous = current
    return previous[-1]


def similarity(a: str, b: str) -> float:
    na = normalize_name(a)
    nb = normalize_name(b)
    if not na or not nb:
        return 0.0
    distance = edit_distance(na, nb)
    return 1 - distance / max(len(na), len(nb))


def is_initial_token(token: str) -> bool:
    return len(token) == 1


def ordered_subsequence(needle: list[str], haystack: list[str]) -> bool:
    if not needle:
        return False
    index = 0
    for token in haystack:
        if token == needle[index]:
            index += 1
            if index == len(needle):
                return True
    return False


def initials_match(short_tokens: list[str], long_tokens: list[str]) -> bool:
    significant = [token for token in short_tokens if not is_initial_token(token)]
    initials = [token for token in short_tokens if is_initial_token(token)]
    if not significant or not initials:
        return False
    if not all(token in long_tokens for token in significant):
        return False
    remaining = [token for token in long_tokens if token not in significant]
    return all(any(candidate.startswith(initial) for candidate in remaining) for initial in initials)


def parse_stage_number(label: Any, fallback: int) -> int:
    if isinstance(label, str):
        match = re.match(r"^\s*(\d+)", label)
        if match:
            return int(match.group(1))
    return fallback


def coerce_rank(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def time_to_seconds(value: time) -> int:
    return value.hour * 3600 + value.minute * 60 + value.second


def seconds_to_display(seconds: int) -> str:
    hours, remainder = divmod(seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours}:{minutes:02d}:{secs:02d}"
    return f"{minutes}:{secs:02d}"


def lookup_standard_stage_distance(stage_number: int) -> int | None:
    return STANDARD_STAGE_DISTANCES_M.get(stage_number)


def parse_time_value(value: Any) -> tuple[str | None, int | None]:
    if value is None or value == "":
        return None, None
    if isinstance(value, time):
        seconds = time_to_seconds(value)
        return seconds_to_display(seconds), seconds
    if isinstance(value, (int, float)):
        seconds = round(float(value) * 24 * 3600)
        return seconds_to_display(seconds), seconds
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None, None
        parts = text.split(":")
        try:
            if len(parts) == 2:
                minutes, secs = int(parts[0]), int(parts[1])
                seconds = minutes * 60 + secs
                return text, seconds
            if len(parts) == 3:
                hours, minutes, secs = int(parts[0]), int(parts[1]), int(parts[2])
                seconds = hours * 3600 + minutes * 60 + secs
                return text, seconds
        except ValueError:
            return text, None
        return text, None
    return str(value), None


def infer_class_code(team_name: str, fallback: Any = None) -> str:
    if isinstance(fallback, str) and fallback.strip():
        return fallback.strip()
    lower = team_name.lower()
    if "veteran" in lower:
        return "Veteran"
    if "miks" in lower or "mix" in lower:
        return "MiksOSI" if "osi" in lower else "MiksSKV"
    if "osi" in lower:
        return "StudOSI"
    if "senior" in lower or "villa" in lower or "queens" in lower:
        return "SeniorSKV" if "sk vidar" in lower or "vidar" in lower else "StudOSI"
    return "EliteSKV"


def infer_organization_code(team_name: str, class_code: str) -> str:
    if class_code in CLASS_META:
        return CLASS_META[class_code]["organization"]
    return "OSIF" if "osi" in team_name.lower() else "SKV"


def choose_stage_meta(
    lookup: dict[tuple[int, str, int, str], dict[str, str]],
    year: int,
    class_code: str,
    stage_number: int,
    raw_name: str,
    fallback: str,
) -> dict[str, str]:
    return lookup.get(
        (year, class_code, stage_number, normalize_name(raw_name)),
        {"label": fallback, "division": "women" if class_code == "Veteran" else "men"},
    )


def parse_split_stage_lookup(workbook_path: Path) -> dict[tuple[int, str, int, str], dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    lookup: dict[tuple[int, str, int, str], dict[str, str]] = {}
    for sheet_name, division in (("HKS_menn_splitt", "men"), ("HKS_kvinner_splitt", "women")):
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            stage_label, raw_name, _, class_code, _, _, year, *_ = row
            if not stage_label or not raw_name or not class_code or not year:
                continue
            stage_number = parse_stage_number(stage_label, 0)
            lookup[(int(year), str(class_code).strip(), stage_number, normalize_name(str(raw_name).strip()))] = {
                "label": str(stage_label).strip(),
                "division": division,
            }
    return lookup


def parse_record_lookup(workbook_path: Path) -> dict[tuple[str, str], dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["Rekorder"]
    lookup: dict[tuple[str, str], dict[str, Any]] = {}
    columns = {
        "men": {"stage": 1, "time": 2, "name": 3, "club": 4, "year": 5},
        "women": {"stage": 8, "time": 9, "name": 10, "club": 11, "year": 12},
    }
    for row_index in range(3, sheet.max_row + 1):
        for division, cols in columns.items():
            stage_value = sheet.cell(row_index, cols["stage"]).value
            time_value = sheet.cell(row_index, cols["time"]).value
            if not stage_value or not time_value:
                continue
            time_text, time_seconds = parse_time_value(time_value)
            lookup[(division, str(stage_value).strip())] = {
                "stage_label": str(stage_value).strip(),
                "record_text": time_text,
                "record_seconds": time_seconds,
                "record_holder": sheet.cell(row_index, cols["name"]).value,
                "record_club": sheet.cell(row_index, cols["club"]).value,
                "record_year": sheet.cell(row_index, cols["year"]).value,
            }
    return lookup


def parse_year_sheets(
    workbook_path: Path, stage_lookup: dict[tuple[int, str, int, str], dict[str, str]]
) -> tuple[list[TeamRecord], list[ResultRecord]]:
    workbook = load_workbook(workbook_path, data_only=True)
    teams: list[TeamRecord] = []
    results: list[ResultRecord] = []
    osif_configs = {
        2025: [
            {"start_col": 2, "class_code": "StudOSI"},
            {"start_col": 6, "class_code": "StudOSI"},
            {"start_col": 10, "class_code": "MiksOSI"},
        ],
        2024: [
            {"start_col": 2, "class_code": "StudOSI"},
            {"start_col": 6, "class_code": "StudOSI"},
        ],
        2023: [
            {"start_col": 2, "class_code": "StudOSI"},
            {"start_col": 6, "class_code": "StudOSI"},
        ],
        2022: [
            {"start_col": 2, "class_code": "StudOSI"},
            {"start_col": 6, "class_code": "StudOSI"},
        ],
    }

    year_sheets = sorted(
        (sheet for sheet in workbook.sheetnames if sheet.isdigit()),
        key=lambda value: int(value),
    )

    for sheet_name in year_sheets:
        year = int(sheet_name)
        sheet = workbook[sheet_name]
        header_rows = [
            row_index
            for row_index in range(1, sheet.max_row + 1)
            if sheet.cell(row_index, 1).value == "Etappe"
        ]
        for header_row in header_rows:
            data_row = header_row + 1
            while data_row <= sheet.max_row and sheet.cell(data_row, 1).value:
                data_row += 1
            total_row = data_row

            if header_row == 38 and year in osif_configs:
                group_index = 0
                for config in osif_configs[year]:
                    start_col = config["start_col"]
                    team_name_value = sheet.cell(header_row, start_col).value
                    if not team_name_value:
                        continue
                    team_name = str(team_name_value).strip()
                    class_code = config["class_code"]
                    group_index += 1
                    total_time_text, total_seconds = parse_time_value(
                        sheet.cell(total_row, start_col + 1).value
                    )
                    teams.append(
                        TeamRecord(
                            year=year,
                            organization_code="OSIF",
                            class_code=class_code,
                            team_name=team_name,
                            source_sheet=sheet_name,
                            header_row=header_row,
                            group_index=group_index,
                            total_time_text=total_time_text,
                            total_seconds=total_seconds,
                            team_rank=coerce_rank(sheet.cell(total_row, start_col + 3).value),
                        )
                    )
                    for offset, row_index in enumerate(range(header_row + 1, total_row), start=1):
                        raw_name_value = sheet.cell(row_index, start_col).value
                        if not raw_name_value:
                            continue
                        stage_label_source = str(sheet.cell(row_index, 1).value).strip()
                        stage_number = parse_stage_number(stage_label_source, offset)
                        stage_meta = choose_stage_meta(
                            stage_lookup,
                            year,
                            class_code,
                            stage_number,
                            str(raw_name_value).strip(),
                            stage_label_source,
                        )
                        split_text, split_seconds = parse_time_value(
                            sheet.cell(row_index, start_col + 1).value
                        )
                        results.append(
                            ResultRecord(
                                year=year,
                                organization_code="OSIF",
                                class_code=class_code,
                                team_name=team_name,
                                source_sheet=sheet_name,
                                source_row=row_index,
                                source_col=start_col,
                                header_row=header_row,
                                group_index=group_index,
                                division=stage_meta["division"],
                                stage_number=stage_number,
                                stage_label=stage_meta["label"],
                                stage_label_source=stage_label_source,
                                raw_name=str(raw_name_value).strip(),
                                split_text=split_text,
                                split_seconds=split_seconds,
                                oa_rank=coerce_rank(sheet.cell(row_index, start_col + 2).value),
                                category_rank=coerce_rank(sheet.cell(row_index, start_col + 3).value),
                            )
                        )
                continue

            group_index = 0
            for start_col in range(2, sheet.max_column + 1, 4):
                team_name_value = sheet.cell(header_row, start_col).value
                if not team_name_value:
                    continue
                team_name = str(team_name_value).strip()
                total_time_text, total_seconds = parse_time_value(
                    sheet.cell(total_row, start_col + 1).value
                )
                class_value = sheet.cell(total_row, start_col + 2).value
                class_code = infer_class_code(team_name, class_value)
                organization_code = infer_organization_code(team_name, class_code)
                group_index += 1
                teams.append(
                    TeamRecord(
                        year=year,
                        organization_code=organization_code,
                        class_code=class_code,
                        team_name=team_name,
                        source_sheet=sheet_name,
                        header_row=header_row,
                        group_index=group_index,
                        total_time_text=total_time_text,
                        total_seconds=total_seconds,
                        team_rank=coerce_rank(sheet.cell(total_row, start_col + 3).value),
                    )
                )
                for offset, row_index in enumerate(range(header_row + 1, total_row), start=1):
                    raw_name_value = sheet.cell(row_index, start_col).value
                    if not raw_name_value:
                        continue
                    stage_label_source = str(sheet.cell(row_index, 1).value).strip()
                    stage_number = parse_stage_number(stage_label_source, offset)
                    stage_meta = choose_stage_meta(
                        stage_lookup,
                        year,
                        class_code,
                        stage_number,
                        str(raw_name_value).strip(),
                        stage_label_source,
                    )
                    split_text, split_seconds = parse_time_value(
                        sheet.cell(row_index, start_col + 1).value
                    )
                    results.append(
                        ResultRecord(
                            year=year,
                            organization_code=organization_code,
                            class_code=class_code,
                            team_name=team_name,
                            source_sheet=sheet_name,
                            source_row=row_index,
                            source_col=start_col,
                            header_row=header_row,
                            group_index=group_index,
                            division=stage_meta["division"],
                            stage_number=stage_number,
                            stage_label=stage_meta["label"],
                            stage_label_source=stage_label_source,
                            raw_name=str(raw_name_value).strip(),
                            split_text=split_text,
                            split_seconds=split_seconds,
                            oa_rank=coerce_rank(sheet.cell(row_index, start_col + 2).value),
                            category_rank=coerce_rank(sheet.cell(row_index, start_col + 3).value),
                        )
                    )
    return teams, results


def load_existing_review_rows() -> dict[tuple[str, str], dict[str, str]]:
    if not REVIEW_PATH.exists():
        return {}
    existing: dict[tuple[str, str], dict[str, str]] = {}
    with REVIEW_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            raw_name = (row.get("raw_name") or "").strip()
            target_name = (row.get("suggested_canonical_name") or "").strip()
            if not raw_name or not target_name:
                continue
            existing[(raw_name, target_name)] = {
                "decision": (row.get("decision") or "").strip(),
                "notes": (row.get("notes") or "").strip(),
            }
    return existing


def build_match_suggestions(results: list[ResultRecord]) -> list[dict[str, Any]]:
    counts = Counter(result.raw_name for result in results)
    unique_names = sorted(counts)
    suggestions: list[dict[str, Any]] = []

    def preferred_name(candidates: list[str]) -> str:
        return sorted(
            candidates,
            key=lambda name: (
                -len(tokenize_name(name)),
                -len(normalize_name(name)),
                -counts[name],
                name,
            ),
        )[0]

    for raw_name in unique_names:
        tokens = tokenize_name(raw_name)
        if len(tokens) < 2:
            continue

        best_candidate: str | None = None
        best_score = 0.0
        best_reason = ""

        for candidate in unique_names:
            if candidate == raw_name:
                continue
            candidate_tokens = tokenize_name(candidate)
            if len(candidate_tokens) < 2:
                continue

            score = similarity(raw_name, candidate)
            same_first = tokens[0] == candidate_tokens[0]
            same_last = tokens[-1] == candidate_tokens[-1]
            first_distance = edit_distance(tokens[0], candidate_tokens[0])
            last_distance = edit_distance(tokens[-1], candidate_tokens[-1])
            subset = set(tokens).issubset(set(candidate_tokens)) or set(candidate_tokens).issubset(set(tokens))

            if normalize_name(raw_name) == normalize_name(candidate):
                score = 0.99
                reason = "Lik stavemåte etter normalisering"
            elif same_first and same_last and subset:
                score = max(score, 0.87)
                reason = "Samme fornavn/etternavn, ulik bruk av mellomnavn"
            elif ordered_subsequence(tokens, candidate_tokens) and len(candidate_tokens) > len(tokens):
                score = max(score, 0.86)
                reason = "Kortform av fullt navn"
            elif initials_match(tokens, candidate_tokens) and len(candidate_tokens) > len(tokens):
                score = max(score, 0.86)
                reason = "Kortform med initialer som matcher fullt navn"
            elif same_first and same_last and score >= 0.75:
                score = max(score, 0.83)
                reason = "Samme fornavn og etternavn med liten skriveforskjell"
            elif (
                same_first
                and len(tokens) == 2
                and len(candidate_tokens) == 2
                and last_distance <= 2
                and similarity(tokens[-1], candidate_tokens[-1]) >= 0.7
            ):
                score = max(score, 0.82)
                reason = "Samme fornavn og liten skriveforskjell i etternavn"
            elif same_last and score >= 0.94 and first_distance <= 2:
                score = max(score, 0.8)
                reason = "Nesten identisk navn med liten staveforskjell"
            else:
                continue

            if score > best_score:
                best_score = score
                best_candidate = candidate
                best_reason = reason

        if best_candidate and best_score >= 0.8:
            chosen = preferred_name([raw_name, best_candidate])
            if chosen == raw_name:
                continue
            suggestions.append(
                {
                    "raw_name": raw_name,
                    "suggested_canonical_name": chosen,
                    "confidence": round(best_score, 2),
                    "reason": best_reason,
                }
            )

    unique_suggestions: dict[str, dict[str, Any]] = {}
    for suggestion in suggestions:
        current = unique_suggestions.get(suggestion["raw_name"])
        if current is None or suggestion["confidence"] > current["confidence"]:
            unique_suggestions[suggestion["raw_name"]] = suggestion
    return sorted(unique_suggestions.values(), key=lambda item: (-item["confidence"], item["raw_name"]))


def write_review_file(suggestions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    existing = load_existing_review_rows()
    rows: list[dict[str, Any]] = []
    for suggestion in suggestions:
        key = (suggestion["raw_name"], suggestion["suggested_canonical_name"])
        prior = existing.get(key, {})
        rows.append(
            {
                "raw_name": suggestion["raw_name"],
                "suggested_canonical_name": suggestion["suggested_canonical_name"],
                "confidence": f'{suggestion["confidence"]:.2f}',
                "reason": suggestion["reason"],
                "decision": prior.get("decision", ""),
                "notes": prior.get("notes", ""),
            }
        )

    for (raw_name, target_name), prior in existing.items():
        if any(
            row["raw_name"] == raw_name and row["suggested_canonical_name"] == target_name for row in rows
        ):
            continue
        rows.append(
            {
                "raw_name": raw_name,
                "suggested_canonical_name": target_name,
                "confidence": "",
                "reason": "Bevart tidligere vurdering",
                "decision": prior.get("decision", ""),
                "notes": prior.get("notes", ""),
            }
        )

    rows.sort(
        key=lambda row: (
            row["decision"].strip().lower() in {"approve", "approved", "reject", "rejected"},
            -float(row["confidence"] or 0),
            row["raw_name"],
        )
    )

    with REVIEW_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "raw_name",
                "suggested_canonical_name",
                "confidence",
                "reason",
                "decision",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    return rows


def create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        PRAGMA foreign_keys = ON;

        CREATE TABLE organizations (
            code TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            short_name TEXT NOT NULL
        );

        CREATE TABLE team_classes (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            official_label TEXT NOT NULL,
            organization_code TEXT NOT NULL REFERENCES organizations(code),
            sort_order INTEGER NOT NULL
        );

        CREATE TABLE events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slug TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL
        );

        CREATE TABLE people (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            canonical_name TEXT NOT NULL UNIQUE,
            normalized_name TEXT NOT NULL
        );

        CREATE TABLE person_aliases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alias_name TEXT NOT NULL UNIQUE,
            normalized_name TEXT NOT NULL,
            person_id INTEGER NOT NULL REFERENCES people(id),
            review_status TEXT NOT NULL DEFAULT 'seed'
        );

        CREATE TABLE teams (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_id INTEGER NOT NULL REFERENCES events(id),
            year INTEGER NOT NULL,
            organization_code TEXT NOT NULL REFERENCES organizations(code),
            class_code TEXT NOT NULL REFERENCES team_classes(code),
            team_name TEXT NOT NULL,
            source_sheet TEXT NOT NULL,
            header_row INTEGER NOT NULL,
            group_index INTEGER NOT NULL,
            total_time_text TEXT,
            total_seconds INTEGER,
            team_rank INTEGER,
            UNIQUE(year, source_sheet, header_row, group_index)
        );

        CREATE TABLE stages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_id INTEGER NOT NULL REFERENCES events(id),
            year INTEGER NOT NULL,
            class_code TEXT NOT NULL REFERENCES team_classes(code),
            stage_number INTEGER NOT NULL,
            stage_label TEXT NOT NULL,
            UNIQUE(year, class_code, stage_number)
        );

        CREATE TABLE results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team_id INTEGER NOT NULL REFERENCES teams(id),
            stage_id INTEGER NOT NULL REFERENCES stages(id),
            raw_name TEXT NOT NULL,
            alias_id INTEGER NOT NULL REFERENCES person_aliases(id),
            split_text TEXT,
            split_seconds INTEGER,
            oa_rank INTEGER,
            category_rank INTEGER,
            source_sheet TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            source_col INTEGER NOT NULL
        );
        """
    )


def build_database(
    teams: list[TeamRecord], results: list[ResultRecord], review_rows: list[dict[str, Any]]
) -> None:
    if DB_PATH.exists():
        DB_PATH.unlink()

    approved_map = build_approved_name_map(review_rows)

    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    create_schema(connection)

    connection.execute(
        "INSERT INTO events (slug, name) VALUES (?, ?)",
        ("holmenkollstafetten", "Holmenkollstafetten"),
    )
    event_id = connection.execute("SELECT id FROM events WHERE slug = ?", ("holmenkollstafetten",)).fetchone()[0]

    for code, meta in ORGANIZATIONS.items():
        connection.execute(
            "INSERT INTO organizations (code, name, short_name) VALUES (?, ?, ?)",
            (code, meta["name"], meta["short_name"]),
        )
    for code, meta in CLASS_META.items():
        connection.execute(
            """
            INSERT INTO team_classes (code, label, official_label, organization_code, sort_order)
            VALUES (?, ?, ?, ?, ?)
            """,
            (code, meta["label"], meta["official_label"], meta["organization"], meta["sort_order"]),
        )

    raw_names = sorted({result.raw_name for result in results})
    canonical_names = sorted(
        {approved_map.get(raw_name, raw_name) for raw_name in raw_names} | set(approved_map.values())
    )
    person_ids: dict[str, int] = {}
    alias_ids: dict[str, int] = {}

    for canonical_name in canonical_names:
        cursor = connection.execute(
            "INSERT INTO people (canonical_name, normalized_name) VALUES (?, ?)",
            (canonical_name, normalize_name(canonical_name)),
        )
        person_ids[canonical_name] = int(cursor.lastrowid)

    for raw_name in raw_names:
        canonical_name = approved_map.get(raw_name, raw_name)
        cursor = connection.execute(
            """
            INSERT INTO person_aliases (alias_name, normalized_name, person_id, review_status)
            VALUES (?, ?, ?, ?)
            """,
            (
                raw_name,
                normalize_name(raw_name),
                person_ids[canonical_name],
                "approved" if raw_name in approved_map else "seed",
            ),
        )
        alias_ids[raw_name] = int(cursor.lastrowid)

    stage_ids: dict[tuple[int, str, int], int] = {}
    for result in sorted(results, key=lambda item: (item.year, item.class_code, item.stage_number)):
        key = (result.year, result.class_code, result.stage_number)
        if key in stage_ids:
            continue
        cursor = connection.execute(
            """
            INSERT INTO stages (event_id, year, class_code, stage_number, stage_label)
            VALUES (?, ?, ?, ?, ?)
            """,
            (event_id, result.year, result.class_code, result.stage_number, result.stage_label),
        )
        stage_ids[key] = int(cursor.lastrowid)

    team_ids: dict[tuple[int, str, int, int], int] = {}
    for team in teams:
        cursor = connection.execute(
            """
            INSERT INTO teams (
                event_id, year, organization_code, class_code, team_name, source_sheet,
                header_row, group_index, total_time_text, total_seconds, team_rank
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                event_id,
                team.year,
                team.organization_code,
                team.class_code,
                team.team_name,
                team.source_sheet,
                team.header_row,
                team.group_index,
                team.total_time_text,
                team.total_seconds,
                team.team_rank,
            ),
        )
        team_ids[(team.year, team.source_sheet, team.header_row, team.group_index)] = int(cursor.lastrowid)

    for result in results:
        team_key = (result.year, result.source_sheet, result.header_row, result.group_index)
        stage_key = (result.year, result.class_code, result.stage_number)
        connection.execute(
            """
            INSERT INTO results (
                team_id, stage_id, raw_name, alias_id, split_text, split_seconds,
                oa_rank, category_rank, source_sheet, source_row, source_col
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                team_ids[team_key],
                stage_ids[stage_key],
                result.raw_name,
                alias_ids[result.raw_name],
                result.split_text,
                result.split_seconds,
                result.oa_rank,
                result.category_rank,
                result.source_sheet,
                result.source_row,
                result.source_col,
            ),
        )

    connection.commit()
    connection.close()


def build_approved_name_map(review_rows: list[dict[str, Any]]) -> dict[str, str]:
    approved_map: dict[str, str] = {}
    for row in review_rows:
        raw_name = row["raw_name"].strip()
        target = row["suggested_canonical_name"].strip()
        decision = row["decision"].strip().lower()
        if raw_name and target and decision in {"approve", "approved"}:
            approved_map[raw_name] = target
    return approved_map


def build_stage_honours(
    results: list[ResultRecord],
    review_rows: list[dict[str, Any]],
    record_lookup: dict[tuple[str, str], dict[str, Any]],
) -> list[dict[str, Any]]:
    approved_map = build_approved_name_map(review_rows)
    grouped: dict[tuple[str, str, int, str], list[dict[str, Any]]] = defaultdict(list)

    for result in results:
        if result.split_seconds is None:
            continue
        record = record_lookup.get((result.division, result.stage_label), {})
        percent_of_record = None
        if record.get("record_seconds"):
            percent_of_record = round(record["record_seconds"] / result.split_seconds * 100, 1)
        grouped[(result.organization_code, result.division, result.stage_number, result.stage_label)].append(
            {
                "person_name": approved_map.get(result.raw_name, result.raw_name),
                "raw_name": result.raw_name,
                "split_text": result.split_text,
                "split_seconds": result.split_seconds,
                "percent_of_record": percent_of_record,
                "class_code": result.class_code,
                "class_label": CLASS_META[result.class_code]["label"],
                "team_name": result.team_name,
                "year": result.year,
                "oa_rank": result.oa_rank,
                "category_rank": result.category_rank,
            }
        )

    group_specs = [
        {
            "key": "skv-men",
            "title": "SK Vidar menn",
            "subtitle": "Topp 5 per etappe på tvers av elite, senior og veteran.",
            "organization_code": "SKV",
            "division": "men",
            "limit": 5,
            "class_codes": {"EliteSKV", "SeniorSKV", "Veteran"},
        },
        {
            "key": "skv-women",
            "title": "SK Vidar kvinner",
            "subtitle": "Topp 5 per etappe på tvers av elite, senior og veteran.",
            "organization_code": "SKV",
            "division": "women",
            "limit": 5,
            "class_codes": {"EliteSKV", "SeniorSKV", "Veteran"},
        },
        {
            "key": "osi-men",
            "title": "OSI Friidrett menn student",
            "subtitle": "Topp 3 per etappe for studentklassen.",
            "organization_code": "OSIF",
            "division": "men",
            "limit": 3,
            "class_codes": {"StudOSI"},
        },
        {
            "key": "osi-women",
            "title": "OSI Friidrett kvinner student",
            "subtitle": "Topp 3 per etappe for studentklassen.",
            "organization_code": "OSIF",
            "division": "women",
            "limit": 3,
            "class_codes": {"StudOSI"},
        },
    ]

    honour_groups: list[dict[str, Any]] = []
    for spec in group_specs:
        stages: list[dict[str, Any]] = []
        matching_keys = sorted(
            [
                key
                for key in grouped
                if key[0] == spec["organization_code"]
                and key[1] == spec["division"]
                and any(
                    entry["class_code"] in spec["class_codes"]
                    for entry in grouped[key]
                )
            ],
            key=lambda item: (item[2], item[3]),
        )
        for _, _, stage_number, stage_label in matching_keys:
            eligible_entries = [
                entry
                for entry in grouped[(spec["organization_code"], spec["division"], stage_number, stage_label)]
                if entry["class_code"] in spec["class_codes"]
            ]
            entries = sorted(
                eligible_entries,
                key=lambda item: (
                    item["split_seconds"],
                    item["category_rank"] if item["category_rank"] is not None else 9999,
                    item["oa_rank"] if item["oa_rank"] is not None else 9999,
                    item["year"],
                ),
            )[: spec["limit"]]
            stages.append(
                {
                    "stage_number": stage_number,
                    "stage_label": stage_label,
                    "distance_m": lookup_standard_stage_distance(stage_number),
                    "record": record_lookup.get((spec["division"], stage_label)),
                    "entries": [
                        {
                            "rank": index + 1,
                            **entry,
                        }
                        for index, entry in enumerate(entries)
                    ],
                }
            )
        honour_groups.append(
            {
                "key": spec["key"],
                "title": spec["title"],
                "subtitle": spec["subtitle"],
                "organization_code": spec["organization_code"],
                "division": spec["division"],
                "limit": spec["limit"],
                "stages": stages,
            }
        )

    return honour_groups


def export_site_data(
    review_rows: list[dict[str, Any]],
    raw_results: list[ResultRecord],
    record_lookup: dict[tuple[str, str], dict[str, Any]],
) -> None:
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row

    results = [
        dict(row)
        for row in connection.execute(
            """
            SELECT
                r.id,
                t.year,
                o.name AS organization_name,
                o.short_name AS organization_code,
                tc.code AS class_code,
                tc.label AS class_label,
                tc.official_label,
                t.team_name,
                t.total_time_text AS team_total_time,
                t.team_rank,
                s.stage_number,
                s.stage_label,
                p.canonical_name AS person_name,
                r.raw_name,
                r.split_text,
                r.split_seconds,
                r.oa_rank,
                r.category_rank
            FROM results r
            JOIN teams t ON t.id = r.team_id
            JOIN organizations o ON o.code = t.organization_code
            JOIN team_classes tc ON tc.code = t.class_code
            JOIN stages s ON s.id = r.stage_id
            JOIN person_aliases pa ON pa.id = r.alias_id
            JOIN people p ON p.id = pa.person_id
            ORDER BY t.year DESC, t.team_name, s.stage_number
            """
        )
    ]

    teams = [
        dict(row)
        for row in connection.execute(
            """
            SELECT
                t.id,
                t.year,
                o.name AS organization_name,
                o.short_name AS organization_code,
                tc.code AS class_code,
                tc.label AS class_label,
                tc.official_label,
                t.team_name,
                t.total_time_text,
                t.total_seconds,
                t.team_rank
            FROM teams t
            JOIN organizations o ON o.code = t.organization_code
            JOIN team_classes tc ON tc.code = t.class_code
            ORDER BY t.year DESC, t.organization_code, tc.sort_order, t.team_name
            """
        )
    ]

    people_rows = [
        dict(row)
        for row in connection.execute(
            """
            SELECT
                p.id,
                p.canonical_name,
                COUNT(r.id) AS appearances,
                COUNT(DISTINCT t.year) AS seasons,
                COUNT(DISTINCT t.team_name) AS teams,
                MIN(r.category_rank) AS best_category_rank,
                SUM(CASE WHEN r.category_rank = 1 THEN 1 ELSE 0 END) AS category_wins,
                MIN(r.split_seconds) AS best_split_seconds
            FROM people p
            LEFT JOIN person_aliases pa ON pa.person_id = p.id
            LEFT JOIN results r ON r.alias_id = pa.id
            LEFT JOIN teams t ON t.id = r.team_id
            GROUP BY p.id
            ORDER BY appearances DESC, category_wins DESC, canonical_name
            """
        )
    ]

    person_teams = defaultdict(set)
    person_classes = defaultdict(set)
    person_years = defaultdict(set)
    person_orgs = defaultdict(set)
    for row in results:
        person_teams[row["person_name"]].add(f'{row["year"]} · {row["team_name"]}')
        person_classes[row["person_name"]].add(row["class_label"])
        person_years[row["person_name"]].add(row["year"])
        person_orgs[row["person_name"]].add(row["organization_code"])

    for row in people_rows:
        row["best_split_text"] = (
            seconds_to_display(row["best_split_seconds"]) if row["best_split_seconds"] is not None else None
        )
        row["classes"] = sorted(person_classes[row["canonical_name"]])
        row["years"] = sorted(person_years[row["canonical_name"]], reverse=True)
        row["organizations"] = sorted(person_orgs[row["canonical_name"]])
        row["team_history"] = sorted(person_teams[row["canonical_name"]], reverse=True)

    class_breakdown = [
        dict(row)
        for row in connection.execute(
            """
            SELECT
                tc.label AS class_label,
                tc.code AS class_code,
                COUNT(r.id) AS result_count,
                COUNT(DISTINCT p.id) AS people_count,
                COUNT(DISTINCT t.id) AS team_count
            FROM team_classes tc
            LEFT JOIN teams t ON t.class_code = tc.code
            LEFT JOIN results r ON r.team_id = t.id
            LEFT JOIN person_aliases pa ON pa.id = r.alias_id
            LEFT JOIN people p ON p.id = pa.person_id
            GROUP BY tc.code
            ORDER BY tc.sort_order
            """
        )
    ]

    year_breakdown = [
        dict(row)
        for row in connection.execute(
            """
            SELECT
                t.year,
                COUNT(r.id) AS result_count,
                COUNT(DISTINCT p.id) AS people_count,
                COUNT(DISTINCT t.id) AS team_count
            FROM teams t
            LEFT JOIN results r ON r.team_id = t.id
            LEFT JOIN person_aliases pa ON pa.id = r.alias_id
            LEFT JOIN people p ON p.id = pa.person_id
            GROUP BY t.year
            ORDER BY t.year DESC
            """
        )
    ]

    pending_reviews = sum(
        1
        for row in review_rows
        if row["decision"].strip().lower() not in {"approve", "approved", "reject", "rejected"}
    )
    stage_honours = build_stage_honours(raw_results, review_rows, record_lookup)

    site_data = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "metadata": {
            "title": "HKSstatsSKV&OSIF",
            "stageDistanceSource": STAGE_DISTANCE_SOURCE,
            "years": sorted({row["year"] for row in teams}, reverse=True),
            "organizations": [
                {"code": code, "name": meta["name"], "shortName": meta["short_name"]}
                for code, meta in ORGANIZATIONS.items()
            ],
            "classes": [
                {
                    "code": code,
                    "label": meta["label"],
                    "officialLabel": meta["official_label"],
                    "organizationCode": meta["organization"],
                }
                for code, meta in CLASS_META.items()
            ],
            "totals": {
                "results": len(results),
                "teams": len(teams),
                "people": len(people_rows),
                "pendingNameReviews": pending_reviews,
            },
        },
        "overview": {
            "kpis": [
                {"label": "Etapper i databasen", "value": len(results)},
                {"label": "Lag importert", "value": len(teams)},
                {"label": "Unike personer", "value": len(people_rows)},
                {"label": "Navn til gjennomgang", "value": pending_reviews},
            ],
            "classBreakdown": class_breakdown,
            "yearBreakdown": year_breakdown,
            "topAppearances": people_rows[:12],
            "topCategoryWins": sorted(
                people_rows,
                key=lambda row: (-row["category_wins"], -row["appearances"], row["canonical_name"]),
            )[:12],
        },
        "stageHonours": stage_honours,
        "results": results,
        "teams": teams,
        "people": people_rows,
        "nameReview": review_rows,
    }

    with (PUBLIC_DATA_DIR / "site-data.json").open("w", encoding="utf-8") as handle:
        json.dump(site_data, handle, ensure_ascii=False, indent=2)

    shutil.copy2(DB_PATH, DOWNLOADS_DIR / DB_PATH.name)
    shutil.copy2(REVIEW_PATH, DOWNLOADS_DIR / REVIEW_PATH.name)
    connection.close()


def main() -> None:
    ensure_directories()
    workbook_path = find_workbook()
    stage_lookup = parse_split_stage_lookup(workbook_path)
    record_lookup = parse_record_lookup(workbook_path)
    teams, results = parse_year_sheets(workbook_path, stage_lookup)
    suggestions = build_match_suggestions(results)
    review_rows = write_review_file(suggestions)
    build_database(teams, results, review_rows)
    export_site_data(review_rows, results, record_lookup)
    print(
        f"Imported {len(results)} etapper, {len(teams)} lag og skrev {DB_PATH.name}, "
        f"site-data.json og {REVIEW_PATH.name}."
    )


if __name__ == "__main__":
    main()
